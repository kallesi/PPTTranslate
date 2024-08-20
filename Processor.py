import zipfile
import xml.etree.ElementTree as ET
import openpyxl
import os
from enum import Enum
from playwright.sync_api import sync_playwright
from constants import *

class Paths(Enum):
  FILE = 0
  FOLDER = 1
  XL = 2
  OUT = 3

class Processor:
  def __init__(self) -> None:
    self._file_path = ""
    self._folder_path = ""
    self._xl_path = ""
    self._out_path = ""

  def set_path(self, file_path: str):
    self._file_path = file_path
    self._folder_path = os.path.dirname(file_path)
    self._xl_path = self._folder_path + r"\Translate.xlsx"
    self._out_path = file_path.replace(".pptx", "_OUT.pptx")
    return self
  
  def get_path(self, path:int):
    if path == Paths.FILE:
      return self._file_path
    if path == Paths.FOLDER:
      return self._folder_path
    if path == Paths.XL:
      return self._xl_path
    if path == Paths.OUT:
      return self._out_path
    
  def extract_text(self):
    # Open the .pptx file as a zip archive
    z = zipfile.ZipFile(self._file_path, "r")
    # Get a list of all slide files
    slide_files = [name for name in z.namelist() if "ppt/slides/slide" in name and not "-rels" in name]
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    # Loop over each slide
    for slide_file in slide_files:
      # Read the XML data
      slide_xml = z.read(slide_file)
      # Parse the xml data
      root = ET.fromstring(slide_xml)
      # Define the namespaces
      namespaces = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                    "p": "http://schemas.openxmlformats.org/presentationml/2006/main"}
      # Loop over each text elem
      for elem in root.findall(".//a:t", namespaces):
        ws.append([elem.text])
      wb.save(self._xl_path)
    z.close()
    return self

  def google_translate(self, source_lang: str = "en", to_lang: str = "zh-CN", css_selector: str = ".ryNqvb",
                       chrome_path: str = CHROME_PATH) -> None:
    """ Translates an excel column into another language using playwright
    :param css_selector: CSS className for the translation results in case Google updates its website structure
    :param chrome_path: Path to your local Chrome installation
    """
    file_path = self._xl_path
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    result = ""
    for row in sheet.iter_rows(min_col=1, max_col=1):
      result += str(row[0].value) + "\n"
    with sync_playwright() as p:
      browser = p.chromium.launch(headless=True, executable_path=chrome_path)
      page = browser.new_page()
      page.goto(f"https://translate.google.com/?sl={source_lang}&tl={to_lang}&op=translate")
      textarea = page.wait_for_selector("//textarea")
      textarea.fill(result)
      page.wait_for_selector(css_selector)
      spans = page.query_selector_all(css_selector)
      translation = ""
      for span in spans:
        translation += span.inner_text()
      browser.close()
      rows = translation.split("\n")
      for i, row in enumerate(rows, start=1):
        sheet.cell(row=i, column=2, value=row)
      wb.save(file_path)

    
  def replace_text(self):
    zin = zipfile.ZipFile(self._file_path, "r")
    zout = zipfile.ZipFile(self._out_path, "w")
    slide_files = [name for name in zin.namelist() if "ppt/slides/slide" in name and not "-rels" in name]
    for item in zin.infolist():
      if item.filename not in slide_files:
        zout.writestr(item, zin.read(item.filename))
    wb = openpyxl.load_workbook(self._xl_path)
    ws = wb.active
    translations = [row[1].value for row in ws.iter_rows()]
    # Loop over each file
    for slide_file in slide_files:
      slide_xml = zin.read(slide_file)
      # Parse xml data
      root = ET.fromstring(slide_xml)
      namespaces = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                    "p": "http://schemas.openxmlformats.org/presentationml/2006/main"}
      for i, elem in enumerate(root.findall(".//a:t", namespaces)):
        if translations[i] == None:
          elem.text = ""
        else:
          elem.text = str(translations[i])
      # Write modified xml back to slide
      slide_xml_new = ET.tostring(root).decode()
      zout.writestr(slide_file, slide_xml_new)
    zin.close()
    zout.close()
      

